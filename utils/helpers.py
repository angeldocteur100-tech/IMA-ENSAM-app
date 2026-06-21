"""
Fonctions utilitaires — IMA Application
"""

import io
import os
from datetime import datetime, timedelta


def format_duration(seconds: float) -> str:
    """Convertit des secondes en format HH:MM:SS lisible."""
    if seconds is None or seconds < 0:
        return "—"
    seconds = int(seconds)
    h, rem = divmod(seconds, 3600)
    m, s   = divmod(rem, 60)
    if h:
        return f"{h:02d}h {m:02d}min {s:02d}s"
    if m:
        return f"{m:02d}min {s:02d}s"
    return f"{s}s"


def format_date(dt_str: str) -> str:
    """Formate une date ISO en format français."""
    if not dt_str:
        return "—"
    try:
        dt = datetime.fromisoformat(dt_str.replace("Z", ""))
        return dt.strftime("%d/%m/%Y")
    except Exception:
        return dt_str


def format_datetime(dt_str: str) -> str:
    """Formate un datetime ISO en format français."""
    if not dt_str:
        return "—"
    try:
        dt = datetime.fromisoformat(dt_str.replace("Z", ""))
        return dt.strftime("%d/%m/%Y à %H:%M")
    except Exception:
        return dt_str


def export_to_excel(data: list[dict], sheet_name: str = "Données") -> bytes:
    """Exporte une liste de dictionnaires vers un fichier Excel en mémoire."""
    try:
        import pandas as pd
        df = pd.DataFrame(data)
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        return buf.getvalue()
    except ImportError:
        return b""


def export_chrono_to_excel(chronos: list[dict]) -> bytes:
    """Exporte les chronométrages vers Excel avec mise en forme."""
    try:
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Alignment

        rows = []
        for c in chronos:
            rows.append({
                "ID":              c.get("id"),
                "Stagiaire":       c.get("stagiaire_nom", ""),
                "Département":     c.get("departement_nom", ""),
                "Mise en situation": c.get("mes_nom", ""),
                "Opération":       c.get("operation_nom", ""),
                "Date":            c.get("date_realisation", ""),
                "Heure":           c.get("heure_realisation", ""),
                "Temps (secondes)": c.get("temps_secondes", 0),
                "Temps (lisible)": format_duration(c.get("temps_secondes", 0)),
                "Saisie manuelle": "Oui" if c.get("saisie_manuelle") else "Non",
            })
        df = pd.DataFrame(rows)
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Chronométrages", index=False)
            ws = writer.sheets["Chronométrages"]
            header_fill = PatternFill("solid", fgColor="1E3A5F")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 20
        return buf.getvalue()
    except ImportError:
        return export_to_excel(chronos, "Chronométrages")


def ensure_upload_dirs():
    """Crée les répertoires d'upload si nécessaire."""
    dirs = [
        "uploads/plans",
        "uploads/fiches_instruction",
        "uploads/fiches_autocontrole",
    ]
    for d in dirs:
        os.makedirs(d, exist_ok=True)


def save_uploaded_file(uploaded_file, folder: str) -> tuple[str, str]:
    """Sauvegarde un fichier uploadé et retourne (nom_fichier, chemin)."""
    ensure_upload_dirs()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = f"{ts}_{uploaded_file.name.replace(' ', '_')}"
    path = os.path.join(folder, safe_name)
    with open(path, "wb") as f:
        f.write(uploaded_file.getbuffer())
    return safe_name, path


def get_role_badge_color(role: str) -> str:
    """Retourne la couleur CSS pour un badge de rôle."""
    colors = {
        "Administrateur": "#dc3545",
        "Formateur":      "#0d6efd",
        "Stagiaire":      "#198754",
    }
    return colors.get(role, "#6c757d")


def get_status_label(statut: int) -> str:
    return "✅ Actif" if statut else "❌ Inactif"
